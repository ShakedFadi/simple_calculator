# __author__ = 'knedlus'
##############
# Calculator #
##############

# Imports

import datetime # For time logging
# import time
# import csv
# import sys
import os # For checking files & for clearing screen in terminal
# import logging
# import xlwt
# import json
import openpyxl # For writing and reading .xlsx log file
#from openpyxl.cell import get_column_letter
#import smtplib
#from email.mime.text import MIMEText
#import sendgrid
import yagmail

# Functions ############################################################################

def add(x, y):
    return x+y

def substract(x, y):
    return x-y

def divide(x, y):
    return x/y

def multiply(x, y):
    return x*y

def clear():
    os.system("clear")
    """
    #print sys.platform
    is_windows = sys.platform.startswith("win")
    if is_windows':
        os.system("cls")
    """
    if os.name == "nt":
        os.system("cls")

# Loggins ############################################################################

start = []
correct_option_selects_time = []
correct_option_selects_option = []
correct_option_selects_result = []
end = []

# Input ############################################################################

def main():
    s = datetime.datetime.now()
    start.append(s)

    clear()
    print "Welcome to Super Simple Ninja Calc!\nPlease choose your operation."
    while True:
        try:
            user_input = int(raw_input("""\n
    1. Add two numbers
    2. Substract two numbers
    3. Divide two numbers
    4. Multiply two numbers\n
    Option: """))

        except ValueError:
            clear()
            print "Please choose your operation by typing a number in front of the chose option!"
            continue


        if user_input < 1 or user_input > 4:
            clear()
            print "Please choose your operation with one of the given numbers!"

        else:
            t = datetime.datetime.now()
            if user_input == 1:
                o = "Add"
            elif user_input == 2:
                o = "Substract"
            elif user_input == 3:
                o = "Divide"
            elif user_input == 4:
                o = "Multiply"
            while True:

                try:
                    clear()
                    if user_input == 1:
                        print "You chose operation Add."
                    elif user_input == 2:
                        print "You chose operation Substract."
                    elif user_input == 3:
                        print "You chose operation Divide."
                    elif user_input == 4:
                        print "You chose operation Multiply."
                    number_x = float(raw_input("\nPlease enter your first number: "))
                    number_y = float(raw_input("Please enter your second number: "))

                except ValueError:
                    print  "\nYou must enter integer or float number!"
                    continue

                if user_input == 1:
                    print """\nYou chose operation Add.
%.01f + %.01f = %.01f\n""" %(number_x, number_y, add(number_x,number_y))
                    r = "%.01f + %.01f = %.01f" %(number_x, number_y, add(number_x,number_y))
                    break

                elif user_input == 2:
                    print """\nYou chose operation Substract.
%.01f - %.01f = %.01f\n""" %(number_x, number_y, substract(number_x,number_y))
                    r = "%.01f - %.01f = %.01f" %(number_x, number_y, substract(number_x,number_y))
                    break

                elif user_input == 3:
                    print """\nYou chose operation Divide.
%.01f / %.01f = %.01f\n""" %(number_x, number_y, divide(number_x,number_y))
                    r = "%.01f / %.01f = %.01f" %(number_x, number_y, divide(number_x,number_y))
                    break

                elif user_input == 4:
                    print """\nYou chose operation Multiply.
%.01f * %.01f = %.01f\n""" %(number_x, number_y, multiply(number_x,number_y))
                    r = "%.01f * %.01f = %.01f" %(number_x, number_y, multiply(number_x,number_y))
                    break

            correct_option_selects_time.append(t)
            correct_option_selects_option.append(o)
            correct_option_selects_result.append(r)

            while True:
                proceed = raw_input("""\nDo you want to perfrom another operation?
Type 'y' for yes or 'n' for no: """).upper()

                if proceed == "Y":
                    clear()
                    print "Please choose your new operation."
                    break

                elif proceed == "N":
                    clear()
                    print "Thank you for using our calculator!\n"
                    return

                else:
                    clear()
                    print "You chose an invalid option!\nTry again!"
                    continue

if __name__ == "__main__":
    main()

e = datetime.datetime.now()
end.append(e)

# Excel logging ############################################################################


exists = os.path.isfile("calculator_loggings.xlsx")

if exists:
    log_file = "calculator_loggings.xlsx"
    wb = openpyxl.load_workbook(filename=log_file)
else:
    wb = openpyxl.Workbook()
    log_file = "calculator_loggings.xlsx"

ws1 = wb.active
ws1.title = "loggings"

ws1["A1"] = "Start"
row_count = ws1.get_highest_row() + 1
ws1["A"+str(row_count)] = start[0]

header=1
ct=2
co=3
cr=4

for t,o,r in zip(correct_option_selects_time,correct_option_selects_option, correct_option_selects_result):
    ws1.cell(row=header, column=(ct+correct_option_selects_time.index(t))).value = "Correct Option Time %s" %(correct_option_selects_time.index(t)+1)
    ws1.cell(row=header, column=(co+correct_option_selects_option.index(o))).value = "Correct Option Name %s" %(correct_option_selects_option.index(o)+1)
    ws1.cell(row=header, column=(cr+correct_option_selects_result.index(r))).value = "Correct Option Result %s" %(correct_option_selects_result.index(r)+1)
    ws1.cell(row=row_count, column=(ct+correct_option_selects_time.index(t))).value = t
    ws1.cell(row=row_count, column=(co+correct_option_selects_option.index(o))).value = o
    ws1.cell(row=row_count, column=(cr+correct_option_selects_result.index(r))).value = r
    ct+=2
    co+=2
    cr+=2


last = ws1.get_highest_column()
lastL = openpyxl.cell.get_column_letter(last)
column_count = ws1.get_highest_column() + 1
needed_col = openpyxl.cell.get_column_letter(column_count)

if ws1[lastL+str(1)].value == "End":
    ws1[lastL+str(row_count)] = end[0]
else:
    ws1[needed_col+str(1)] = "End"
    ws1[needed_col+str(row_count)] = end[0]


wb.save(filename=log_file)
print correct_option_selects_time
print len(correct_option_selects_time)
# Sending Email with logging ############################################################################

def send_loggings(u,p,st):

    loggings_file = "calculator_loggings.xlsx"
    body = """Oi we are sendign you attachment with Calculator Loggings, created by Ninja Calc!
Check the data! :)"""

    yag = yagmail.SMTP(u, p)
    yag.send(to = st, subject ="Calculator Loggings on %s" %datetime.datetime.now().strftime("%A, %d.%m.%Y, %H:%M:%S"), contents = [body, loggings_file])


while True:

    send_option = raw_input("""Do you want your loggings to be send to an email address of your choice?
    Type 'y' for yes or 'n' for no: """).upper()

    if send_option == "Y":
        clear()
        mf = raw_input("Please enter email address you want your mail sent from: ")
        mp = raw_input("Please enter password relative to email address above: ")
        mt = raw_input("Finally please enter email addres you want your mail sent to: ")

        send_loggings(mf, mp, mt)
        clear()
        print """Loggings of your chose calculator operations and history have been sent to listed email address!
Have a nice day!\n"""
        break


    elif send_option == "N":
        clear()
        print """Thank you for using our calculator!
Your results and loggings have been saved in calculator_loggings.xlsx file in simple_calculator foleder.
Bye!\n"""
        break

    else:
        clear()
        print "You chose an invalid option!\nTry again!\n"
        continue

